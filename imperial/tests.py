import io
import zipfile
from unittest.mock import patch

from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, TransactionTestCase, override_settings
from PIL import Image

from imperial.models import Apartamento, Leitura, PortalConfig
from imperial.storage import SupabaseStorage


PORTAL_URL = "/imperial"
DOWNLOAD_PHOTOS_URL = "/dw-imperial"
DOWNLOAD_EXCEL_URL = "/imperial/excel"


def make_png_bytes():
    buf = io.BytesIO()
    Image.new("RGB", (1, 1), color="red").save(buf, format="PNG")
    return buf.getvalue()


class ImperialPortalGetTests(TestCase):
    def setUp(self):
        Apartamento.objects.create(apartamento="101")
        Apartamento.objects.create(apartamento="202")

    def test_get_portal_aberto_renderiza_apartamentos(self):
        config = PortalConfig.get_solo()
        config.is_open = True
        config.save()

        response = self.client.get(PORTAL_URL)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["is_portal_open"])
        self.assertEqual(len(response.context["apartamentos"]), 2)

    def test_get_portal_fechado_marca_flag_no_contexto(self):
        config = PortalConfig.get_solo()
        config.is_open = False
        config.save()

        response = self.client.get(PORTAL_URL)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["is_portal_open"])


class ImperialPortalPostTests(TestCase):
    def setUp(self):
        self.apartamento = Apartamento.objects.create(apartamento="101")
        config = PortalConfig.get_solo()
        config.is_open = True
        config.save()

    def _post_payload(self, **overrides):
        payload = {
            "apartamento": str(self.apartamento.id),
            "valor_leitura": "1234,567",
            "foto_relogio": SimpleUploadedFile(
                "leitura.png", make_png_bytes(), content_type="image/png"
            ),
        }
        payload.update(overrides)
        return payload

    @patch.object(SupabaseStorage, "_save", return_value="fake/path.png")
    def test_post_sucesso_cria_leitura(self, _mock_save):
        response = self.client.post(PORTAL_URL, self._post_payload())

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Leitura.objects.count(), 1)
        leitura = Leitura.objects.first()
        self.assertEqual(leitura.apartamento, self.apartamento)
        self.assertEqual(str(leitura.valor_leitura), "1234.567")

        messages = [m.message for m in response.context["messages"]]
        self.assertTrue(any("sucesso" in m.lower() for m in messages))

    @patch.object(SupabaseStorage, "_save", return_value="fake/path.png")
    def test_post_sem_apartamento_mostra_erro(self, _mock_save):
        payload = self._post_payload()
        payload["apartamento"] = ""

        response = self.client.post(PORTAL_URL, payload)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Leitura.objects.count(), 0)
        messages = [m.message for m in response.context["messages"]]
        self.assertTrue(any("erro" in m.lower() for m in messages))

    def test_post_portal_fechado_mostra_warning(self):
        config = PortalConfig.get_solo()
        config.is_open = False
        config.save()

        response = self.client.post(PORTAL_URL, self._post_payload())

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Leitura.objects.count(), 0)
        messages = [m.message for m in response.context["messages"]]
        self.assertTrue(any("fechado" in m.lower() for m in messages))


class ImperialPortalDuplicateTests(TransactionTestCase):
    def setUp(self):
        self.apartamento = Apartamento.objects.create(apartamento="101")
        config = PortalConfig.get_solo()
        config.is_open = True
        config.save()

    @patch.object(
        SupabaseStorage,
        "_save",
        side_effect=ValidationError("duplicate_upload"),
    )
    def test_post_duplicado_mostra_warning_especifico(self, _mock_save):
        payload = {
            "apartamento": str(self.apartamento.id),
            "valor_leitura": "1234,567",
            "foto_relogio": SimpleUploadedFile(
                "leitura.png", make_png_bytes(), content_type="image/png"
            ),
        }
        response = self.client.post(PORTAL_URL, payload)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Leitura.objects.count(), 0)
        messages = [m.message for m in response.context["messages"]]
        self.assertTrue(any("já enviou" in m for m in messages))


@override_settings(MEDIA_ROOT="/tmp/leitura-nova-tests-imperial")
class ImperialDownloadPhotosTests(TestCase):
    def test_download_photos_retorna_zip_vazio_quando_sem_fotos(self):
        response = self.client.get(DOWNLOAD_PHOTOS_URL)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        self.assertIn("photos_imperial.zip", response["Content-Disposition"])

        zf = zipfile.ZipFile(io.BytesIO(response.content))
        self.assertEqual(zf.namelist(), [])


class ImperialDownloadExcelTests(TestCase):
    def setUp(self):
        self.apto1 = Apartamento.objects.create(apartamento="101")
        self.apto2 = Apartamento.objects.create(apartamento="202")
        Leitura.objects.create(
            apartamento=self.apto1,
            data_leitura="2026-05-01",
            valor_leitura="100.500",
        )

    def test_download_excel_retorna_xlsx_com_dados(self):
        response = self.client.get(DOWNLOAD_EXCEL_URL)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("leituras_imperial.xlsx", response["Content-Disposition"])

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = [tuple(cell.value for cell in row) for row in ws.iter_rows()]

        self.assertEqual(rows[0], ("Apartamento", "Data Leitura", "Valor Leitura"))
        apartamentos_nas_linhas = {row[0] for row in rows[1:]}
        self.assertEqual(apartamentos_nas_linhas, {"101", "202"})
